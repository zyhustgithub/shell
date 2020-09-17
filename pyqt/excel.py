from openpyxl import load_workbook
'''
from enum import Enum
class DATABASE(Enum):
    PRODUCT_CATEGORY    = 0
    PRODUCT_NAME        = 1
    PRODUCT_SEND_PATH   = 2
    PRODUCT_SEND_NAME   = 3
    IF_UZIP             = 4
    RECORD_LENGTH       = 5
print(DATABASE.PRODUCT_CATEGORY, DATABASE.PRODUCT_CATEGORY.name, DATABASE.PRODUCT_CATEGORY.value)
'''
PRODUCT_CATEGORY    = 0
PRODUCT_NAME        = 1
PRODUCT_SEND_PATH   = 2
PRODUCT_SEND_NAME   = 3
IF_UZIP             = 4

RECORD_LENGTH       = 5 #字段数

def read_excel(cat=None, file_path='products.xlsx'):
    workbook = load_workbook(file_path)

    #booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    sheets = workbook.get_sheet_names()         #从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows
    #columns = booksheet.columns
    content = []
    #breack_line = 0
    #end_line = booksheet.max_row
    head = True
    for row in rows: #迭代所有的行
        record = [col.value for col in row[:RECORD_LENGTH]]
        if not (record[PRODUCT_CATEGORY] and record[PRODUCT_CATEGORY].strip() and \
                record[PRODUCT_NAME] and record[PRODUCT_NAME].strip() and \
                record[PRODUCT_SEND_PATH] and record[PRODUCT_SEND_PATH].strip()): #如果没有产品类别或者名称或者发送路径则视为无效数据，并退出循环
            break
        tmp = []
        for v in record:
            if v and v.strip():
                tmp.append(v.strip())
            else:
                tmp.append(None)
        record = tmp
        if head == True:
            content.append(record)
            head = False
            continue
        if not cat:
            if record not in content:
                content.append(record)
            continue
        if cat in "".join([rec for rec in record if rec]):
            #breack_line = row[0].row
            #print(breack_line)
            if record not in content:
                content.append(record)
    tmp = list(["".join(rec[0:3]) for rec in content])
    if len(tmp) != len(set(tmp)):
        content = []
    del tmp
    #print(content)
    #print(content[breack_line-1])
    return content

if __name__ == '__main__':
    read_excel('清算后')
    read_excel('第一批')